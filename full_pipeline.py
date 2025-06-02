# full_pipeline.py (HTML Dashboard Version - No More Excel Drama!)

import os
from scrape_goodreads_backlist import search_goodreads_author, scrape_goodreads_books
import pandas as pd
from openpyxl import load_workbook
import time
import re
from html import escape

# ----------------------- SCRAPE PHASE -----------------------
print("[1/2] Scraping Goodreads backlists...")

# Load authors from your real convention xlsx
wb = load_workbook("announced_authors.xlsx")
ws = wb.active

data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    author = row[0]
    role = row[1]
    other_names = row[2]
    website_link = row[3]
    goodreads_link = row[4]
    amazon_link = row[5]
    audible_link = row[6]
    
    data.append({
        "Author Name": author,
        "Role": role,
        "Other Names": other_names,
        "Website": website_link,
        "Goodreads Page": goodreads_link,
        "Amazon Page": amazon_link,
        "Audible Page": audible_link
    })

author_df = pd.DataFrame(data)

# Load existing scraped data if it exists
if os.path.exists("author_backlists_scraped.xlsx"):
    existing_data = pd.read_excel("author_backlists_scraped.xlsx", engine="openpyxl")
    scraped_authors = existing_data["Author"].dropna().unique().tolist()
    print(f"Found existing scraped data. {len(scraped_authors)} authors already scraped.")
else:
    existing_data = pd.DataFrame()
    scraped_authors = []

# Determine which entries still need to be scraped
entries_to_scrape = [entry for entry in data if entry["Author Name"] not in scraped_authors]
print(f"Entries to scrape: {[e['Author Name'] for e in entries_to_scrape]}")

new_books = []
for idx, row in author_df.iterrows():
    name = str(row.get("Author Name", "")).strip()
    if not name or name.lower() == "nan":
        print(f"⚠️  Skipping row {idx} — missing Author Name")
        continue

    if name in scraped_authors:
        continue  # already scraped

    role = str(row.get("Role", "")).strip() or "Author"
    other_names_raw = row.get("Other Names")
    if pd.isna(other_names_raw):
        other_names_raw = ""
    pen_names = [n.strip() for n in str(other_names_raw).split(",") if n.strip()]

    names_to_scrape = [name] + pen_names

    # Get the author data for later use
    author_row = None
    for entry in data:
        if entry["Author Name"] == name:
            author_row = entry
            break

    for pen_name in names_to_scrape:
        print(f"🔍 Scraping {pen_name} for {name} ({role})...")
        author_url = search_goodreads_author(pen_name)
        if author_url:
            books = scrape_goodreads_books(author_url, name, role, pen_name)
            for book in books:
                book["Author"] = name
                book["Pen Name"] = pen_name if pen_name != name else ""
                book["Role"] = role
            new_books.extend(books)
        time.sleep(2)
    print(f"✅ Finished scraping {name}")

# Merge new data with existing data
if new_books:
    new_data = pd.DataFrame(new_books)
    full_data = pd.concat([existing_data, new_data], ignore_index=True)
else:
    full_data = existing_data

# Save updated data
full_data.to_excel("author_backlists_scraped.xlsx", index=False)
print("Scraping complete. Data saved to author_backlists_scraped.xlsx\n")

# ----------------------- HTML DASHBOARD PHASE -----------------------
print("[2/2] Building HTML dashboard...")

def clean_url(url):
    """Clean and validate URL"""
    if not url or pd.isna(url) or str(url).strip() == "":
        return None
    
    url = str(url).strip()
    if url.startswith("http://") or url.startswith("https://"):
        return url
    else:
        return "https://" + url

def create_html_dashboard():
    # All helper functions consolidated here
    def clean_field(field_value):
        if pd.isna(field_value) or str(field_value).strip() in ['', 'nan', 'None']:
            return ""
        return escape(str(field_value).strip())
    
    def parse_series_from_title(title_text):
        """
        Parse series info from book title patterns like:
        - "Book Title (Series Name, #1)"
        - "Book Title (Series Name #1)" 
        - "Book Title (Series Name Book 1)"
        Returns: (clean_title, series_name, series_order)
        """
        if not title_text or pd.isna(title_text):
            return "", "", ""
        
        title = str(title_text).strip()
        
        # Pattern 1: (Series Name, #1) or (Series Name, Book 1)
        pattern1 = r'^(.*?)\s*\(\s*([^,]+),\s*(?:#|Book\s*)(\d+)\s*\)'
        match1 = re.match(pattern1, title, re.IGNORECASE)
        if match1:
            clean_title = match1.group(1).strip()
            series_name = match1.group(2).strip()
            series_order = match1.group(3).strip()
            return clean_title, series_name, series_order
        
        # Pattern 2: (Series Name #1) or (Series Name Book 1)
        pattern2 = r'^(.*?)\s*\(\s*([^#]+?)(?:\s*#|Book\s*)(\d+)\s*\)'
        match2 = re.match(pattern2, title, re.IGNORECASE)
        if match2:
            clean_title = match2.group(1).strip()
            series_name = match2.group(2).strip()
            series_order = match2.group(3).strip()
            return clean_title, series_name, series_order
        
        # Pattern 3: (Series Name) - assume book 1 or standalone
        pattern3 = r'^(.*?)\s*\(\s*([^)]+)\s*\)'
        match3 = re.match(pattern3, title, re.IGNORECASE)
        if match3:
            clean_title = match3.group(1).strip()
            series_name = match3.group(2).strip()
            # Only treat as series if it contains certain keywords
            if any(keyword in series_name.lower() for keyword in ['series', 'saga', 'chronicles', 'trilogy', 'duology']):
                return clean_title, series_name, "1"
            else:
                # Might be series name, but we're not sure of order
                return clean_title, series_name, ""
        
        # No series info found in title, return as standalone
        return title, "", ""
    
    def format_yes_no_maybe(field_value):
        clean_val = clean_field(field_value).lower()
        if clean_val in ['yes', 'y', 'true', '1']:
            return '<span class="yes-no-cell yes-cell">Yes</span>'
        elif clean_val in ['no', 'n', 'false', '0']:
            return '<span class="yes-no-cell no-cell">No</span>'
        elif clean_val in ['maybe', 'm', '?', 'possible']:
            return '<span class="yes-no-cell" style="color: #ffc107; font-weight: bold;">Maybe</span>'
        elif clean_val:
            return f'<span class="yes-no-cell">{escape(str(field_value))}</span>'
        else:
            return '<span class="yes-no-cell">-</span>'
    
    def determine_standalone_series(series_title, series_order):
        """Determine if book is standalone or part of series"""
        if series_title and series_title.strip() and series_title.lower() != "nan":
            return "Series"
        else:
            return "Standalone"
    
    def determine_audiobook_status(author_name, role, author_data):
        """Determine audiobook availability based on role and Audible presence"""
        # All narrators get "Yes"
        if role and role.lower() == "narrator":
            return "Yes"
        
        # Authors with Audible links get "Maybe"
        if author_data and author_data.get("Audible Page"):
            audible_url = str(author_data.get("Audible Page", "")).strip()
            if audible_url and audible_url.lower() not in ["", "nan", "none"]:
                return "Maybe"
        
        # No Audible link = "No"
        return "No"

    html_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Charm City Romanticon 2026 - Author Backlists</title>
    <style>
        * {
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            line-height: 1.6;
        }
        
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            overflow: hidden;
        }
        
        .header {
            background: #EC008C;
            color: white;
            padding: 40px 30px;
            text-align: center;
        }
        
        .header h1 {
            margin: 0;
            font-size: 2.8em;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        }
        
        .header p {
            margin: 10px 0 0 0;
            font-size: 1.2em;
            opacity: 0.9;
        }
        
        .support-message {
            background: linear-gradient(45deg, #f8f9fa, #e9ecef);
            padding: 25px;
            margin: 25px;
            border-radius: 12px;
            border-left: 6px solid #EC008C;
            font-style: italic;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        
        .support-message strong {
            color: #EC008C;
            font-size: 1.1em;
        }
        
        .search-bar {
            padding: 20px;
            text-align: center;
            background: #f8f9fa;
            border-bottom: 1px solid #eee;
            display: flex;
            justify-content: center;
            align-items: center;
            gap: 20px;
            flex-wrap: wrap;
        }
        
        .search-input {
            padding: 12px 20px;
            font-size: 16px;
            border: 2px solid #ddd;
            border-radius: 25px;
            width: 300px;
            max-width: 90%;
            outline: none;
            transition: border-color 0.3s ease;
        }
        
        .search-input:focus {
            border-color: #EC008C;
        }
        
        .export-btn {
            padding: 12px 25px;
            background: linear-gradient(45deg, #28a745, #20c997);
            color: white;
            border: none;
            border-radius: 25px;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
            text-decoration: none;
            display: inline-flex;
            align-items: center;
            gap: 8px;
        }
        
        .export-btn:hover {
            background: linear-gradient(45deg, #218838, #1ea085);
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(40, 167, 69, 0.3);
        }
        
        .authors-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(380px, 1fr));
            gap: 25px;
            padding: 30px;
        }
        
        .author-card {
            background: white;
            border: 2px solid #EC008C;
            border-radius: 15px;
            padding: 25px;
            box-shadow: 0 5px 20px rgba(0,0,0,0.1);
            transition: all 0.3s ease;
            position: relative;
            overflow: hidden;
        }
        
        .author-card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 4px;
            background: linear-gradient(90deg, #EC008C, #ff6b9d);
        }
        
        .author-card:hover {
            transform: translateY(-8px);
            box-shadow: 0 15px 35px rgba(0,0,0,0.15);
            border-color: #d1007a;
        }
        
        .author-name {
            font-size: 1.5em;
            font-weight: bold;
            color: #EC008C;
            margin-bottom: 8px;
        }
        
        .author-role {
            color: #666;
            margin-bottom: 20px;
            font-style: italic;
            font-size: 1.05em;
        }
        
        .links {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
            gap: 10px;
            margin-bottom: 20px;
        }
        
        .link-btn {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            padding: 10px 15px;
            background: linear-gradient(45deg, #EC008C, #ff6b9d);
            color: white;
            text-decoration: none;
            border-radius: 8px;
            transition: all 0.3s ease;
            font-size: 0.9em;
            font-weight: 500;
            text-align: center;
        }
        
        .link-btn:hover {
            background: linear-gradient(45deg, #d1007a, #e55a87);
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(236, 0, 140, 0.3);
        }
        
        .link-btn span {
            margin-right: 8px;
            font-size: 1.1em;
        }
        
        .books-section {
            margin-top: 20px;
            padding-top: 20px;
            border-top: 2px solid #f0f0f0;
        }
        
        .books-toggle {
            background: linear-gradient(45deg, #f8f9fa, #e9ecef);
            border: 2px solid #EC008C;
            padding: 12px 20px;
            border-radius: 8px;
            cursor: pointer;
            font-weight: bold;
            margin-bottom: 15px;
            transition: all 0.3s ease;
            color: #EC008C;
            text-align: center;
        }
        
        .books-toggle:hover {
            background: #EC008C;
            color: white;
        }
        
        .books-list {
            display: none;
            max-height: 400px;
            overflow-y: auto;
            font-size: 0.95em;
        }
        
        .books-list.show {
            display: block;
            animation: slideDown 0.3s ease;
        }
        
        @keyframes slideDown {
            from { opacity: 0; transform: translateY(-10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .books-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
            font-size: 0.85em;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        
        .books-table th {
            background: #EC008C;
            color: white;
            padding: 12px 8px;
            text-align: left;
            font-weight: bold;
            font-size: 0.8em;
            border-bottom: 2px solid #d1007a;
        }
        
        .books-table td {
            padding: 10px 8px;
            border-bottom: 1px solid #f0f0f0;
            vertical-align: top;
        }
        
        .books-table tr:hover {
            background: #f8f9fa;
        }
        
        .books-table tr:nth-child(even) {
            background: #fafafa;
        }
        
        .books-table tr:nth-child(even):hover {
            background: #f0f0f0;
        }
        
        .book-title-cell {
            font-weight: bold;
            color: #333;
            min-width: 150px;
        }
        
        .series-cell {
            color: #666;
            min-width: 120px;
        }
        
        .yes-no-cell {
            text-align: center;
            font-weight: bold;
        }
        
        .yes-cell {
            color: #28a745;
        }
        
        .no-cell {
            color: #dc3545;
        }
        
        .link-cell a {
            color: #EC008C;
            text-decoration: none;
            font-weight: 500;
        }
        
        .link-cell a:hover {
            text-decoration: underline;
        }
        
        .table-container {
            overflow-x: auto;
            margin-top: 15px;
        }
        
        @media (max-width: 768px) {
            .books-table {
                font-size: 100px; /* Force table to be scrollable width */
            }
            
            .books-table th,
            .books-table td {
                padding: 8px 4px;
            }
        }
        
        .footer {
            text-align: center;
            padding: 30px;
            background: #f8f9fa;
            font-style: italic;
            color: #666;
            border-top: 1px solid #eee;
        }
        
        .stats {
            text-align: center;
            padding: 20px;
            background: #f8f9fa;
            color: #666;
            font-size: 0.9em;
        }
        
        .hidden {
            display: none !important;
        }
        
        @media (max-width: 768px) {
            .authors-grid {
                grid-template-columns: 1fr;
                padding: 20px;
                gap: 20px;
            }
            
            .header h1 {
                font-size: 2.2em;
            }
            
            .support-message {
                margin: 15px;
                padding: 20px;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📚 Charm City Romanticon 2026</h1>
            <p>Author & Narrator Backlists</p>
        </div>
        
        <div class="support-message">
            <strong>💡 Support Authors Directly</strong><br>
            Whenever possible, consider purchasing books directly from the author's website if they have a store.
            Amazon takes a significant portion of royalties and can penalize authors for piracy and other issues beyond their control — even removing their accounts.
            We understand that Amazon is convenient and affordable, and authors still rely on it.
            But every direct purchase makes a bigger impact. 💖
        </div>
        
        <div class="disclaimer-message" style="background: linear-gradient(45deg, #ff6b6b, #ff8e8e); padding: 25px; margin: 25px; border-radius: 12px; border-left: 6px solid #dc3545; font-style: italic; box-shadow: 0 2px 10px rgba(0,0,0,0.1); color: white;">
            <strong>⚠️ Amazon Scraping Disclaimer</strong><br>
            My goal was to scrape all this data online and deliver a more comprehensive dashboard with detailed buy links, library availability, narrator info, and more. Unfortunately, Amazon is a massive pain in the ass and doesn't allow you to effectively scrape their data. So a big F*** YOU to Amazon! 🖕 Therefore, I had to condense the columns I would have liked to include. I'm sorry for the limitations, but blame Amazon's anti-scraping fortress, not me! 😤
        </div>
        
        <div class="search-bar">
            <input type="text" class="search-input" placeholder="Search authors..." onkeyup="searchAuthors()">
            <button class="export-btn" onclick="exportToCSV()">
                📊 Export to Excel/Sheets
            </button>
        </div>
        
        <div class="stats" id="stats">
            Loading authors...
        </div>
        
        <div class="authors-grid" id="authorsGrid">
"""
    
    # Track stats
    total_authors = 0
    total_books = 0
    
    # Add each author
    for person in sorted(full_data["Author"].dropna().unique()):
        person_data = full_data[full_data["Author"].str.lower() == person.lower()]
        role = person_data["Role"].iloc[0] if "Role" in person_data else "Author"
        
        # Find author info
        author_row = None
        for entry in data:
            if entry["Author Name"] == person:
                author_row = entry
                break
        
        if not author_row:
            continue
        
        total_authors += 1
        books = person_data.to_dict('records')
        total_books += len(books)
        
        # Clean person name for JavaScript
        clean_person = escape(person).replace("'", "\\'")
        
        html_content += f"""
            <div class="author-card" data-name="{escape(person.lower())}" data-role="{escape(role.lower())}">
                <div class="author-name">{escape(person)}</div>
                <div class="author-role">{escape(role)}</div>
                <div class="links">
        """
        
        # Add links only if they exist
        links_added = 0
        if clean_url(author_row.get("Website")):
            html_content += f'<a href="{clean_url(author_row.get("Website"))}" class="link-btn" target="_blank"><span>🌐</span>Website</a>'
            links_added += 1
        
        if clean_url(author_row.get("Goodreads Page")):
            html_content += f'<a href="{clean_url(author_row.get("Goodreads Page"))}" class="link-btn" target="_blank"><span>📚</span>Goodreads</a>'
            links_added += 1
        
        if clean_url(author_row.get("Amazon Page")):
            html_content += f'<a href="{clean_url(author_row.get("Amazon Page"))}" class="link-btn" target="_blank"><span>🛒</span>Amazon</a>'
            links_added += 1
        
        if clean_url(author_row.get("Audible Page")):
            html_content += f'<a href="{clean_url(author_row.get("Audible Page"))}" class="link-btn" target="_blank"><span>🎧</span>Audible</a>'
            links_added += 1
        
        if links_added == 0:
            html_content += '<div style="text-align: center; color: #999; font-style: italic;">Links coming soon!</div>'
        
        html_content += '</div>'
        
        # Add books section
        if books:
            html_content += f"""
                <div class="books-section">
                    <div class="books-toggle" onclick="toggleBooks('{clean_person}')">
                        📖 View Books ({len(books)})
                    </div>
                    <div id="books-{clean_person}" class="books-list">
                        <div class="table-container">
                            <table class="books-table">
                                <thead>
                                    <tr>
                                        <th>Book Title</th>
                                        <th>Standalone/Series</th>
                                        <th>Series</th>
                                        <th>Order</th>
                                        <th>Published Year</th>
                                        <th>Formats</th>
                                        <th>Audio</th>
                                        <th>Pen Name</th>
                                    </tr>
                                </thead>
                                <tbody>
            """
            
            for book in books:
                # Parse book title for series info
                raw_title = book.get("Book Title", "")
                clean_title, parsed_series, parsed_order = parse_series_from_title(raw_title)
                
                # Use parsed data or fall back to existing data
                title = escape(clean_title) if clean_title else clean_field(raw_title)
                
                # For series, prefer existing data, then parsed data
                existing_series = clean_field(book.get("Series Title", ""))
                series = existing_series if existing_series else parsed_series
                
                # For order, prefer existing data, then parsed data  
                existing_order = clean_field(book.get("Series Order", ""))
                series_order = existing_order if existing_order else parsed_order
                
                # Try multiple date field names and clean up the format
                published_date = ""
                possible_date_fields = [
                    "Published Date", "Release Date", "Publication Date", "Date Published",
                    "Published", "Release", "Publication", "Date", "Pub Date", "Publish Date"
                ]
                
                for date_field in possible_date_fields:
                    if book.get(date_field):
                        raw_date_value = book.get(date_field)
                        
                        # Skip if the value is NaN or None
                        if pd.isna(raw_date_value) or raw_date_value is None:
                            continue
                        
                        # Handle float years (like 2019.0) and convert to clean integers
                        if isinstance(raw_date_value, (int, float)):
                            # Additional check for NaN floats
                            if pd.isna(raw_date_value):
                                continue
                                
                            # If it's a number, assume it's a year
                            try:
                                year = int(raw_date_value)
                                if 1900 <= year <= 2030:  # Reasonable year range
                                    published_date = str(year)
                                    break
                            except (ValueError, OverflowError):
                                continue
                        else:
                            # If it's a string, clean it
                            date_value = clean_field(raw_date_value)
                            if date_value:
                                # Try to extract just the year from string dates
                                year_match = re.search(r'\b(\d{4})\b', date_value)
                                if year_match:
                                    try:
                                        year = int(year_match.group(1))
                                        if 1900 <= year <= 2030:
                                            published_date = str(year)
                                            break
                                    except ValueError:
                                        continue
                                else:
                                    # If no year found, use the original string
                                    published_date = date_value
                                    break
                
                # Other fields
                formats = clean_field(book.get("Formats Available", ""))
                
                # Determine standalone vs series properly
                standalone_series = determine_standalone_series(series, series_order)
                
                # Determine audiobook status based on role and Audible presence
                audiobook_status = determine_audiobook_status(person, role, author_row)
                audiobook = format_yes_no_maybe(audiobook_status)
                
                pen_name = clean_field(book.get("Pen Name", ""))
                
                # Only show pen name if different from main author name
                if pen_name.lower() == person.lower():
                    pen_name = ""
                
                html_content += f"""
                    <tr>
                        <td class="book-title-cell">{title or "-"}</td>
                        <td>{standalone_series}</td>
                        <td class="series-cell">{series or "-"}</td>
                        <td>{series_order or "-"}</td>
                        <td>{published_date or "-"}</td>
                        <td>{formats or "-"}</td>
                        <td>{audiobook}</td>
                        <td>{pen_name or "-"}</td>
                    </tr>
                """
            
            html_content += """
                                </tbody>
                            </table>
                        </div>
                    </div>
                </div>
            """
        
        html_content += "</div>"
    
    html_content += f"""
        </div>
        
        <div class="footer">
            Compiled for Charm City Romanticon 2026 by Plot Twists & Pivot Tables
        </div>
    </div>
    
    <script>
        // Update stats
        document.getElementById('stats').innerHTML = `📊 {total_authors} Authors & Narrators • {total_books} Books`;
        
        function toggleBooks(author) {{
            const booksList = document.getElementById('books-' + author);
            booksList.classList.toggle('show');
        }}
        
        function searchAuthors() {{
            const searchTerm = document.querySelector('.search-input').value.toLowerCase();
            const cards = document.querySelectorAll('.author-card');
            let visibleCount = 0;
            
            cards.forEach(card => {{
                const name = card.dataset.name;
                const role = card.dataset.role;
                const isVisible = name.includes(searchTerm) || role.includes(searchTerm);
                
                if (isVisible) {{
                    card.classList.remove('hidden');
                    visibleCount++;
                }} else {{
                    card.classList.add('hidden');
                }}
            }});
            
            // Update stats
            if (searchTerm) {{
                document.getElementById('stats').innerHTML = `🔍 Showing ${{visibleCount}} results for "${{searchTerm}}"`;
            }} else {{
                document.getElementById('stats').innerHTML = `📊 {total_authors} Authors & Narrators • {total_books} Books`;
            }}
        }}
        
        function exportToCSV() {{
            const csvData = [];
            
            // Add header row
            csvData.push([
                'Author Name',
                'Role', 
                'Book Title',
                'Standalone/Series',
                'Series',
                'Order',
                'Published Year',
                'Formats',
                'Audio',
                'Pen Name'
            ]);
            
            // Get all author cards
            const authorCards = document.querySelectorAll('.author-card');
            
            authorCards.forEach(card => {{
                const authorName = card.querySelector('.author-name').textContent;
                const authorRole = card.querySelector('.author-role').textContent;
                
                // Get the books table for this author
                const booksTable = card.querySelector('.books-table tbody');
                
                if (booksTable) {{
                    const rows = booksTable.querySelectorAll('tr');
                    
                    rows.forEach(row => {{
                        const cells = row.querySelectorAll('td');
                        if (cells.length > 0) {{
                            const rowData = [
                                authorName,
                                authorRole,
                                cells[0]?.textContent?.trim() || '', // Book Title
                                cells[1]?.textContent?.trim() || '', // Standalone/Series
                                cells[2]?.textContent?.trim() || '', // Series
                                cells[3]?.textContent?.trim() || '', // Order
                                cells[4]?.textContent?.trim() || '', // Published Year
                                cells[5]?.textContent?.trim() || '', // Formats
                                cells[6]?.textContent?.trim() || '', // Audio
                                cells[7]?.textContent?.trim() || ''  // Pen Name
                            ];
                            csvData.push(rowData);
                        }}
                    }});
                }} else {{
                    // If no books, add just the author info
                    csvData.push([authorName, authorRole, '', '', '', '', '', '', '', '']);
                }}
            }});
            
            // Convert to CSV format
            const csvContent = csvData.map(row => 
                row.map(cell => {{
                    // Escape quotes and wrap in quotes if contains comma, quote, or newline
                    const escapedCell = String(cell).replace(/"/g, '""');
                    return escapedCell.includes(',') || escapedCell.includes('"') || escapedCell.includes('\\n') 
                        ? `"${{escapedCell}}"` 
                        : escapedCell;
                }}).join(',')
            ).join('\\n');
            
            // Create download
            const blob = new Blob([csvContent], {{ type: 'text/csv;charset=utf-8;' }});
            const link = document.createElement('a');
            
            if (link.download !== undefined) {{
                const url = URL.createObjectURL(blob);
                link.setAttribute('href', url);
                link.setAttribute('download', 'charm_city_romanticon_2026_backlists.csv');
                link.style.visibility = 'hidden';
                document.body.appendChild(link);
                link.click();
                document.body.removeChild(link);
                
                // Show success message
                const btn = document.querySelector('.export-btn');
                const originalText = btn.innerHTML;
                btn.innerHTML = '✅ Downloaded!';
                btn.style.background = 'linear-gradient(45deg, #28a745, #20c997)';
                
                setTimeout(() => {{
                    btn.innerHTML = originalText;
                    btn.style.background = 'linear-gradient(45deg, #28a745, #20c997)';
                }}, 2000);
            }}
        }}
        
        // Add some smooth scrolling
        document.querySelectorAll('a[href^="#"]').forEach(anchor => {{
            anchor.addEventListener('click', function (e) {{
                e.preventDefault();
                document.querySelector(this.getAttribute('href')).scrollIntoView({{
                    behavior: 'smooth'
                }});
            }});
        }});
    </script>
</body>
</html>
    """
    
    # Save the HTML file
    with open("charm_city_romanticon_2026_backlists.html", "w", encoding="utf-8") as f:
        f.write(html_content)
    
    print("✅ HTML Dashboard created: charm_city_romanticon_2026_backlists.html")
    print(f"   📊 {total_authors} authors/narrators with {total_books} total books")
    print("   🌐 Just double-click the file to open in your browser!")
    print("   📱 Works on desktop, tablet, and mobile")

# Create the beautiful HTML dashboard
create_html_dashboard()
print("\n🎉 Done! No more Excel drama - just pure HTML awesomeness!")