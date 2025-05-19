# excel_backlist_builder.py

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re

# Load scraped data
scraped_data = pd.read_excel("author_backlists_scraped.xlsx", engine='openpyxl')

# Initialize workbook
wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

# Create Dashboard sheet
dashboard = wb.create_sheet("Dashboard")
dashboard.append(["Author Name", "Link to Tab"])

# Define styles
hot_pink_fill = PatternFill(start_color="EC008C", end_color="EC008C", fill_type="solid")
black_font_bold = Font(color="000000", bold=True)
gray_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Define headers
headers = [
    "Book Title", "Series Title", "Series Order", "Published Date",
    "Formats Available", "Buy Links", "Rent Links", "Audiobook (Y/N)",
    "Narrators", "Kindle Unlimited (Y/N)", "Kobo+ (Y/N)",
    "Genre", "Standalone/Series", "Other Notes"
]

# Process each unique author
authors = scraped_data["Author"].dropna().unique()
for author in authors:
    author_data = scraped_data[scraped_data["Author"] == author]
    tab_name = re.sub(r'[\\/*?:"<>|]', '', author)  # Clean tab name
    tab_name = author if len(author) <= 31 else author[:28] + "..."
    ws = wb.create_sheet(tab_name)

    # Connect with Author block
    ws.merge_cells('A1:B1')
    ws['A1'] = f"Connect with {author}"
    ws['A1'].alignment = Alignment(horizontal='left')
    ws['A1'].font = black_font_bold
    ws['A1'].fill = hot_pink_fill
    for row in range(1, 5):
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
    ws['A2'] = "🌐 Website"
    ws['B2'] = ""
    ws['A3'] = "📚 Goodreads"
    ws['B3'] = ""
    ws['A4'] = "🛒 Amazon"
    ws['B4'] = ""

    # Spacer row
    ws.append([])
    ws.append([])
    ws.append(headers)

    # Format header
    for col_num, column_title in enumerate(headers, 1):
        cell = ws[f'{get_column_letter(col_num)}7']
        cell.font = black_font_bold
        cell.fill = hot_pink_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Add book rows
    for idx, row_data in enumerate(author_data.itertuples(index=False), start=8):
        row_list = [
            row_data._1, row_data._2, row_data._3, row_data._4,
            row_data._5, "", "", "", "", "", "", "", "", ""
        ]
        ws.append(row_list)
        for col_num in range(1, len(headers)+1):
            cell = ws[f'{get_column_letter(col_num)}{idx}']
            if idx % 2 == 1:
                cell.fill = gray_fill
            cell.border = thin_border
            if col_num == 4:
                cell.number_format = 'MM/DD/YYYY'

    # Add to dashboard
    if [author, f"='{tab_name}'!A1"] not in dashboard:
        dashboard.append([author, f"='{tab_name}'!A1"])

# Style dashboard header
for col in range(1, 3):
    cell = dashboard.cell(row=1, column=col)
    cell.fill = hot_pink_fill
    cell.font = black_font_bold
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Border entire dashboard table
for row_idx in range(1, dashboard.max_row + 1):
    for col_idx in range(1, 3):
        dashboard.cell(row=row_idx, column=col_idx).border = thin_border

# Footer
footer_row = dashboard.max_row + 3
footer_text = "Compiled for Charm City Romanticon 2026 by Plot Twists & Pivot Tables"
dashboard.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=2)
dashboard.cell(row=footer_row, column=1).value = footer_text
dashboard.cell(row=footer_row, column=1).alignment = Alignment(horizontal='center')
dashboard.cell(row=footer_row, column=1).font = Font(italic=True)

# Save file
wb.save("author_backlist_final.xlsx")
print("Excel file created: author_backlist_final.xlsx")
